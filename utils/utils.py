from settings import *
from openpyxl.styles import Font
import openpyxl

def new_module(value_product, caracter_product, href_produts):
    """
        Cria uma planilha Excel com características, valor e link de acesso de um produto.

        Este método cria uma nova planilha Excel contendo as características do produto, o valor do produto e o link de acesso.
        A planilha é salva com um nome de arquivo baseado no valor do produto.

        Args:
            value_product (Any): Valor do produto a ser incluído na planilha.
            caracter_product (List[str]): Lista de características do produto a serem incluídas na planilha.
            href_produts (str): Link de acesso à página do produto a ser incluído na planilha.

        Returns:
            Dict[str, Any]: Um dicionário contendo os seguintes campos:
                - "error" (bool): Indica se ocorreu algum erro durante a execução.
                - "message" (str or None): Mensagem de sucesso ou de erro, se houver.
                - "data" (str or None): Nome do arquivo gerado, se bem-sucedido; caso contrário, None.
    """
    try:
        wb = openpyxl.Workbook()
        ws = wb.active

        ws.title = "CARACTERISTICAS DO PRODUTO"

        ws["A1"] = "DESCRIÇÃO E CARACTERISTICAS"
        ws["A1"].font = Font(bold=True)
        
        ws["B1"] = "VALOR DO PRODUTO"
        ws["B1"].font = Font(bold=True)

        ws["C1"] = "LINK DE ACESSO"
        ws["C1"].font = Font(bold=True)

        ws.cell(row=2, column=2).value = value_product
        ws.cell(row=2, column=3).value = href_produts

        for i, caracter in enumerate(caracter_product, start=2):
            ws.cell(row=i, column=1).value = caracter
        
        wb.save(f"caracteristicas_produto_{value_product}.xlsx")

    except Exception as e:
        logger.error(f"Error: {str(e)}")
        return {"error": True, "message": f"Error: {str(e)}", "data": None}

    return {"error": False, "message": "Planilha gerada com sucesso!", "data": "caracteristicas_produto.xlsx"}
